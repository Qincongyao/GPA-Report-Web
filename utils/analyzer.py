# -*- coding: utf-8 -*-
"""
GPA 成绩分析核心模块
"""

import os
import re
import json
import io
import base64
from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'SimHei', 'Microsoft YaHei']
plt.rcParams['axes.unicode_minus'] = False


# 配置常量
ID_COL_STD = "学号"
CLASS_COL_STD = "班级"
NAME_COL_STD = "姓名"
GPA_COL_STD = "总GPA"
SHEET_NAME = "数据"
BASE_HEADER = [ID_COL_STD, NAME_COL_STD, CLASS_COL_STD]

# 阈值配置
LOW = 3.0
MED = 3.5


class GPAAanalyzer:
    """GPA 成绩分析器"""
    
    def __init__(self, low=LOW, med=MED):
        self.LOW = low
        self.MED = med
        
    def process_file(self, filepath):
        """处理单个文件"""
        # 创建临时汇总表
        temp_xlsx = f"temp_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        
        try:
            # 导入数据
            wb, ws = self._ensure_book_and_sheet(temp_xlsx)
            date_label, _, _, _ = self._date_from_name(filepath)
            df_raw = self._read_source_raw(filepath)
            df_std = self._normalize_source(df_raw)
            self._upsert_wide(ws, date_label, df_std)
            self._sort_rows_and_columns(ws)
            self._apply_conditional_colors(ws)
            wb.save(temp_xlsx)
            
            # 生成HTML
            rows, classes, dates = self._read_wide_sheet(temp_xlsx, SHEET_NAME)
            rows = self._add_change_rate(rows, dates)
            html = self._build_html(rows, classes, dates)
            
            return {
                'html': html,
                'stats': {
                    'studentCount': len(rows),
                    'dateCount': len(dates),
                    'classCount': len(classes),
                    'dates': dates
                }
            }
        finally:
            # 清理临时文件
            if os.path.exists(temp_xlsx):
                os.remove(temp_xlsx)
    
    def process_multiple_files(self, filepaths, lang='zh'):
        temp_xlsx = f"temp_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        
        try:
            wb, ws = self._ensure_book_and_sheet(temp_xlsx)
            
            imported_dates = []
            for filepath in filepaths:
                try:
                    date_label, _, _, _ = self._date_from_name(filepath)
                    df_raw = self._read_source_raw(filepath)
                    df_std = self._normalize_source(df_raw)
                    self._upsert_wide(ws, date_label, df_std)
                    imported_dates.append(date_label)
                except Exception as e:
                    print(f"导入失败 {filepath}: {e}")
            
            self._sort_rows_and_columns(ws)
            self._apply_conditional_colors(ws)
            wb.save(temp_xlsx)
            
            rows, classes, dates = self._read_wide_sheet(temp_xlsx, SHEET_NAME)
            rows = self._add_change_rate(rows, dates)
            html = self._build_html(rows, classes, dates, lang=lang)
            
            return {
                'html': html,
                'stats': {
                    'studentCount': len(rows),
                    'dateCount': len(dates),
                    'classCount': len(classes),
                    'dates': dates
                }
            }
        finally:
            if os.path.exists(temp_xlsx):
                os.remove(temp_xlsx)
    
    def generate_pdf(self, filepaths, lang='zh'):
        """生成PDF报告（统计分析 + 折线图）"""
        temp_xlsx = f"temp_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        
        try:
            wb, ws = self._ensure_book_and_sheet(temp_xlsx)
            
            for filepath in filepaths:
                try:
                    date_label, _, _, _ = self._date_from_name(filepath)
                    df_raw = self._read_source_raw(filepath)
                    df_std = self._normalize_source(df_raw)
                    self._upsert_wide(ws, date_label, df_std)
                except Exception as e:
                    print(f"导入失败 {filepath}: {e}")
            
            self._sort_rows_and_columns(ws)
            wb.save(temp_xlsx)
            
            rows, classes, dates = self._read_wide_sheet(temp_xlsx, SHEET_NAME)
            rows = self._add_change_rate(rows, dates)
            
            # 生成PDF
            pdf_path = f"report_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
            
            with PdfPages(pdf_path) as pdf:
                # 第一页：封面 + 统计
                self._build_pdf_cover_and_stats(pdf, rows, classes, dates, lang)
                
                # 后续页：每个学生的折线图
                self._build_pdf_charts(pdf, rows, dates, classes)
            
            return pdf_path
        finally:
            if os.path.exists(temp_xlsx):
                os.remove(temp_xlsx)
    
    def _build_pdf_cover_and_stats(self, pdf, rows, classes, dates, lang):
        """生成PDF封面和统计页"""
        from matplotlib.backends.backend_pdf import PdfPages
        import matplotlib.pyplot as plt
        
        # ========== 第一页：全年级统计 ==========
        fig1 = plt.figure(figsize=(11, 8.5))
        
        title = 'GPA Analysis Report' if lang == 'en' else 'GPA 成绩分析报告'
        fig1.suptitle(title, fontsize=24, fontweight='bold', y=0.98)
        
        # 计算全年级统计数据（排除0和空值）
        all_stats = []
        for d in dates:
            gpas = []
            for r in rows:
                v = r.get(d)
                if v is not None and v != '' and v != 0:
                    try:
                        gpas.append(float(v))
                    except:
                        pass
            
            if gpas:
                avg = sum(gpas) / len(gpas)
                above_4 = sum(1 for g in gpas if g >= 4.0)
                red_count = sum(1 for g in gpas if g < self.LOW)
                yel_count = sum(1 for g in gpas if self.LOW <= g < self.MED)
                all_stats.append([d, f"{avg:.2f}", above_4, red_count, yel_count])
        
        # 表头
        col_labels = ['Date', 'Average', '>=4.0', 'Red (<3.0)', 'Yellow (3.0-3.5)'] if lang == 'en' else ['日期', '平均分', '>=4.0', '红色预警', '黄色预警']
        
        ax_table = fig1.add_subplot(211)
        ax_table.axis('off')
        
        if all_stats:
            table = ax_table.table(
                cellText=all_stats,
                colLabels=col_labels,
                loc='center',
                cellLoc='center',
                colColours=['#E8E8E8'] * 5
            )
            table.auto_set_font_size(False)
            table.set_fontsize(10)
            table.scale(1.2, 1.8)
        
        # 统计摘要
        ax_summary = fig1.add_subplot(212)
        ax_summary.axis('off')
        
        total_students = len(rows)
        total_classes = len(classes)
        
        # 计算最新日期的统计
        latest_date = dates[-1] if dates else ''
        latest_gpas = []
        for r in rows:
            v = r.get(latest_date)
            if v is not None and v != '' and v != 0:
                try:
                    latest_gpas.append(float(v))
                except:
                    pass
        
        if latest_gpas:
            latest_avg = sum(latest_gpas) / len(latest_gpas)
            summary_text = f"Total: {total_students} students | {total_classes} classes | Latest Avg: {latest_avg:.2f}"
            if lang == 'zh':
                summary_text = f"学生总数: {total_students} | 班级数: {total_classes} | 最新平均分: {latest_avg:.2f}"
        else:
            summary_text = f"Total: {total_students} students | {total_classes} classes"
            if lang == 'zh':
                summary_text = f"学生总数: {total_students} | 班级数: {total_classes}"
        
        ax_summary.text(0.5, 0.5, summary_text, ha='center', va='center', fontsize=14, 
                       transform=ax_summary.transAxes, bbox=dict(boxstyle='round', facecolor='#F0F0F0', alpha=0.8))
        
        plt.tight_layout(rect=[0, 0, 1, 0.95])
        pdf.savefig(fig1)
        plt.close(fig1)
        
        # ========== 第二页起：每个班级的统计 ==========
        # 按班级分组
        rows_by_class = {}
        for r in rows:
            cls = r.get('班级', 'Unknown')
            if cls not in rows_by_class:
                rows_by_class[cls] = []
            rows_by_class[cls].append(r)
        
        for cls in sorted(rows_by_class.keys()):
            class_rows = rows_by_class[cls]
            
            fig2 = plt.figure(figsize=(11, 8.5))
            fig2.suptitle(f"Class {cls}", fontsize=20, fontweight='bold', y=0.98)
            
            ax_table2 = fig2.add_subplot(111)
            ax_table2.axis('off')
            
            # 计算该班级的统计数据
            class_stats = []
            for d in dates:
                gpas = []
                for r in class_rows:
                    v = r.get(d)
                    if v is not None and v != '' and v != 0:
                        try:
                            gpas.append(float(v))
                        except:
                            pass
                
                if gpas:
                    avg = sum(gpas) / len(gpas)
                    above_4 = sum(1 for g in gpas if g >= 4.0)
                    red_count = sum(1 for g in gpas if g < self.LOW)
                    yel_count = sum(1 for g in gpas if self.LOW <= g < self.MED)
                    class_stats.append([d, f"{avg:.2f}", above_4, red_count, yel_count])
            
            if class_stats:
                table2 = ax_table2.table(
                    cellText=class_stats,
                    colLabels=col_labels,
                    loc='center',
                    cellLoc='center',
                    colColours=['#DDEEFF'] * 5
                )
                table2.auto_set_font_size(False)
                table2.set_fontsize(11)
                table2.scale(1.2, 2.0)
            
            plt.tight_layout(rect=[0, 0, 1, 0.95])
            pdf.savefig(fig2)
            plt.close(fig2)
    
    def _build_pdf_charts(self, pdf, rows, dates, classes):
        """生成每个学生的折线图"""
        import matplotlib.pyplot as plt
        from pandas import DataFrame
        
        if not dates or not rows:
            return
        
        df = DataFrame(rows)
        x_index = {d: i for i, d in enumerate(dates)}
        
        # 按班级分组
        rows_by_class = {}
        for r in rows:
            cls = r.get('班级', 'Unknown')
            if cls not in rows_by_class:
                rows_by_class[cls] = []
            rows_by_class[cls].append(r)
        
        # 每个班级一页图表
        for cls in sorted(rows_by_class.keys()):
            class_rows = rows_by_class[cls]
            
            # 每页最多6个学生
            for i in range(0, len(class_rows), 6):
                subset = class_rows[i:i+6]
                fig, axes = plt.subplots(2, 3, figsize=(11, 8.5))
                fig.suptitle(f"Class {cls}", fontsize=16, fontweight='bold')
                
                axes_flat = axes.flatten()
                
                for j, r in enumerate(subset):
                    ax = axes_flat[j]
                    name = r.get('姓名', '')
                    sid = r.get('学号', '')
                    
                    xs = []
                    ys = []
                    for d in dates:
                        v = r.get(d)
                        if v and str(v).replace('.', '').replace('-', '').isdigit():
                            xs.append(x_index[d])
                            ys.append(float(v))
                    
                    if xs:
                        ax.plot(xs, ys, marker='o', linewidth=2, markersize=6)
                        ax.set_ylim(2, 5)
                        ax.set_xticks(range(len(dates)))
                        ax.set_xticklabels(dates, rotation=45, ha='right', fontsize=8)
                        ax.axhline(self.MED, color='#facc15', linestyle='--', linewidth=1, alpha=0.7)
                        ax.axhline(self.LOW, color='#ef4444', linestyle='--', linewidth=1, alpha=0.7)
                        ax.set_title(f"{name} ({sid})", fontsize=10)
                        ax.set_ylabel('GPA', fontsize=8)
                        ax.grid(True, alpha=0.3)
                        
                        # 标注数值
                        for xi, y in zip(xs, ys):
                            ax.annotate(f'{y:.2f}', (xi, y), textcoords="offset points", 
                                       xytext=(0, 5), ha='center', fontsize=7)
                
                # 隐藏多余的子图
                for j in range(len(subset), 6):
                    axes_flat[j].axis('off')
                
                plt.tight_layout(rect=[0, 0, 1, 0.94])
                pdf.savefig(fig)
                plt.close(fig)
    
    def generate_pdf_from_data(self, rows, classes, dates, lang='zh'):
        """从已有数据生成PDF报告（用于筛选导出）"""
        # 添加变化率
        rows = self._add_change_rate(rows, dates)
        
        pdf_path = f"report_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
        
        with PdfPages(pdf_path) as pdf:
            # 第一页：封面 + 统计
            self._build_pdf_cover_and_stats(pdf, rows, classes, dates, lang)
            
            # 后续页：每个学生的折线图
            self._build_pdf_charts(pdf, rows, dates, classes)
        
        return pdf_path
    
    # ========== 数据处理方法 ==========
    
    def _date_from_name(self, path: str) -> str:
        stem = os.path.splitext(os.path.basename(path))[0]
        stem = re.sub(r"\(\d+\)$", "", stem)
        m = re.search(r"(\d{8})", stem)
        if not m:
            raise ValueError(f"无法从文件名解析日期：{stem}")
        yyyymmdd = m.group(1)
        year = int(yyyymmdd[:4])
        month = int(yyyymmdd[4:6])
        day = int(yyyymmdd[6:8])
        return f"{year}/{month:02d}/{day:02d}", year, month, day

    def _parse_mmdd(self, s, default_year=2025):
        """Parse date with year support"""
        # Try format: YYYY/MM/DD
        m = re.fullmatch(r"(\d{4})/(\d{1,2})/(\d{1,2})", str(s).strip()) if s else None
        if m:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        # Try format: MM/DD (assume default year)
        m = re.fullmatch(r"(\d{1,2})/(\d{1,2})", str(s).strip()) if s else None
        return datetime(default_year, int(m.group(1)), int(m.group(2))) if m else None

    def _read_source_raw(self, path: str) -> pd.DataFrame:
        ext = os.path.splitext(path)[1].lower()
        eng = "xlrd" if ext == ".xls" else "openpyxl"
        return pd.read_excel(path, engine=eng)

    def _normalize_class(self, val: str) -> str:
        s = str(val or "").strip()
        m = re.search(r"(\d+)", s)
        if m:
            return f"Class {int(m.group(1))}"
        return s

    def _clean_name(self, s: str) -> str:
        if s is None:
            return ""
        name = str(s).strip()
        name = name.replace("（", "(").replace("）", ")")
        name = re.sub(r"\s+", " ", name)
        
        m = re.search(r"\((.*?)\)", name)
        if m:
            inner = m.group(1).strip()
            first_inner = re.split(r"[\\/、,，]", inner)[0].strip()
            before = name.split("(", 1)[0].strip()
            return f"{before} ({first_inner})" if first_inner else before
        
        if "/" in name:
            left, right = name.split("/", 1)
            left = left.strip()
            right = re.split(r"[\\/、,，]", right.strip())[0].strip()
            return f"{left} ({right})" if right else left
        
        return name

    def _normalize_source(self, df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        df.columns = [str(c).strip() for c in df.columns]

        id_alias = ["学号", "Student No.", "Student No", "Student Number"]
        class_alias = ["班级", "Class"]
        cname_alias = ["中文名", "Chinese Name"]
        ename_alias = ["英文名", "English Name"]
        gpa_alias = ["总GPA", "Overall GPA"]

        def pick(colnames):
            for c in colnames:
                if c in df.columns:
                    return c
            return None

        id_col = pick(id_alias)
        cls_col = pick(class_alias)
        cn_col = pick(cname_alias)
        en_col = pick(ename_alias)
        gpa_col = pick(gpa_alias)

        if not id_col: raise KeyError(f"缺少学号列")
        # 班级列改为可选
        if not gpa_col: raise KeyError(f"缺少GPA列")
        if not (cn_col or en_col): raise KeyError(f"缺少姓名列")

        if cn_col:
            name_series = df[cn_col].astype(str).str.strip()
            if en_col:
                en_series = df[en_col].astype(str).str.strip()
                name_series = name_series.mask(
                    name_series.eq("") | name_series.str.lower().eq("nan"),
                    en_series
                )
        else:
            name_series = df[en_col].astype(str).str.strip()
        name_series = name_series.map(self._clean_name)
        
        # 班级列是可选的
        if cls_col:
            class_data = df[cls_col].astype(str).str.strip()
        else:
            class_data = ["All"] * len(df)
        
        out = pd.DataFrame({
            ID_COL_STD: df[id_col].astype(str).str.strip(),
            CLASS_COL_STD: class_data,
            NAME_COL_STD: name_series.astype(str).str.strip(),
            GPA_COL_STD: pd.to_numeric(df[gpa_col], errors="coerce"),
        })

        out = out.dropna(subset=[ID_COL_STD])
        out = out[out[ID_COL_STD].astype(str).str.strip() != ""]
        out[CLASS_COL_STD] = out[CLASS_COL_STD].map(self._normalize_class)
        
        return out

    def _ensure_book_and_sheet(self, xlsx_path: str):
        wb = load_workbook(xlsx_path) if os.path.exists(xlsx_path) else Workbook()
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.active if len(wb.sheetnames)==1 and wb.active.max_row<=1 else wb.create_sheet(SHEET_NAME, 0)
            ws.title = SHEET_NAME
        
        header = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
        if not header or header[:3] != BASE_HEADER:
            ws.delete_rows(1, ws.max_row)
            for i, v in enumerate(BASE_HEADER, start=1):
                ws.cell(1, i).value = v
        return wb, ws

    def _get_or_create_date_col(self, ws, date_label: str) -> int:
        for c in range(4, ws.max_column+1):
            if ws.cell(1, c).value == date_label:
                return c
        col = ws.max_column + 1
        ws.cell(1, col).value = date_label
        return col

    def _build_row_index(self, ws):
        idx = {}
        for r in range(2, ws.max_row+1):
            sid = ws.cell(r, 1).value
            if sid:
                idx[str(sid).strip()] = r
        return idx

    def _upsert_wide(self, ws, date_label: str, df_std: pd.DataFrame):
        col = self._get_or_create_date_col(ws, date_label)
        idx = self._build_row_index(ws)
        next_row = ws.max_row + 1

        df = df_std.copy()
        df = df.dropna(subset=[ID_COL_STD])
        df[ID_COL_STD] = df[ID_COL_STD].astype(str).str.strip()
        df = df.sort_values([CLASS_COL_STD, NAME_COL_STD], kind="stable")

        for _, row in df.iterrows():
            sid = str(row[ID_COL_STD]).strip()
            name = str(row[NAME_COL_STD]).strip()
            cls = str(row[CLASS_COL_STD]).strip()
            gpa = row[GPA_COL_STD]

            if not sid:
                continue

            if sid in idx:
                r = idx[sid]
            else:
                r = next_row
                next_row += 1
                idx[sid] = r
                ws.cell(r, 1).value = sid
                ws.cell(r, 2).value = name
                ws.cell(r, 3).value = cls

            if name: ws.cell(r, 2).value = name
            if cls: ws.cell(r, 3).value = cls

            if pd.notna(gpa):
                try:
                    ws.cell(r, col).value = float(gpa)
                except:
                    ws.cell(r, col).value = gpa

    def _sort_rows_and_columns(self, ws):
        date_cols = []
        for c in range(4, ws.max_column+1):
            label = ws.cell(1, c).value
            date_cols.append((c, label, self._parse_mmdd(label) or datetime.max))
        date_cols.sort(key=lambda x: x[2])
        target_labels = [lab for _, lab, _ in date_cols]

        data = []
        for r in range(2, ws.max_row+1):
            row = [
                ws.cell(r, 1).value,
                ws.cell(r, 2).value,
                ws.cell(r, 3).value,
            ] + [ws.cell(r, c).value for c, _, _ in date_cols]
            data.append(row)

        ws.delete_rows(1, ws.max_row)
        ws.append(BASE_HEADER + target_labels)

        data.sort(key=lambda x: (str(x[2]) or "", str(x[1]) or ""))
        for row in data:
            ws.append(row)

    def _apply_conditional_colors(self, ws):
        red_fill = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
        yellow_fill = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
        dark_red = Font(color="9C0006")
        dark_yellow = Font(color="9C6500")

        if ws.max_row < 2 or ws.max_column < 4:
            return
        for c in range(4, ws.max_column+1):
            col_letter = get_column_letter(c)
            rng = f"{col_letter}2:{col_letter}{ws.max_row}"
            ws.conditional_formatting.add(
                rng,
                CellIsRule(operator='lessThan', formula=[str(self.MED)],
                           stopIfTrue=False, fill=yellow_fill, font=dark_yellow)
            )
            ws.conditional_formatting.add(
                rng,
                CellIsRule(operator='lessThan', formula=[str(self.LOW)],
                           stopIfTrue=True, fill=red_fill, font=dark_red)
            )

    # ========== HTML生成方法 ==========
    
    def _read_wide_sheet(self, xlsx_path, sheet_name):
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb[sheet_name]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
        
        date_cols = []
        for c in range(4, ws.max_column + 1):
            label = headers[c - 1]
            date_cols.append((c, label))
        date_labels = [lbl for _, lbl in date_cols]
        
        records = []
        classes = set()
        for r in range(2, ws.max_row+1):
            sid = ws.cell(r, 1).value
            name = ws.cell(r, 2).value
            cls = ws.cell(r, 3).value
            if sid is None and name is None and cls is None:
                continue
            row = {
                "学号": str(sid or "").strip(),
                "姓名": (name or "").strip(),
                "班级": self._normalize_class(cls),
            }
            classes.add(row["班级"])
            for c, lbl in date_cols:
                v = ws.cell(r, c).value
                try:
                    row[str(lbl)] = float(v) if v not in [None, ""] else None
                except (TypeError, ValueError):
                    row[str(lbl)] = None
            records.append(row)

        records = [r for r in records if r["学号"]]
        return records, sorted(classes), date_labels

    def _fig_to_base64(self, fig):
        buf = io.BytesIO()
        fig.savefig(buf, format="png", bbox_inches="tight")
        buf.seek(0)
        return base64.b64encode(buf.read()).decode("utf-8")

    def _build_chart_cards(self, df, dates):
        cards = []
        if df.empty or not dates:
            return ""

        long = df.melt(id_vars=["学号", "姓名", "班级"], value_vars=dates,
                       var_name="时间", value_name="总GPA").dropna(subset=["总GPA"])
        if long.empty:
            return ""

        x_index = {d:i for i, d in enumerate(dates)}

        for sid, g in long.groupby("学号", sort=True):
            name = str(g["姓名"].iloc[0]) if "姓名" in g.columns and len(g) else ""
            cls = str(g["班级"].iloc[0]) if "班级" in g.columns and len(g) else ""

            g = g.sort_values("时间")
            xs = [x_index.get(t, None) for t in g["时间"]]
            ys = pd.to_numeric(g["总GPA"], errors="coerce").tolist()

            if all((y is None or pd.isna(y)) for y in ys):
                continue

            has_red = any(isinstance(v, (int, float)) and v != 0 and v < self.LOW for v in ys)
            has_yel = any(isinstance(v, (int, float)) and v != 0 and (self.LOW <= v < self.MED) for v in ys)

            fig, ax = plt.subplots(figsize=(4.0, 2.2), dpi=150)
            ax.plot(xs, ys, linestyle="--", marker="o")
            ax.set_ylim(2, 5)
            ax.set_ylabel("GPA")
            ax.set_xlabel("Date")
            ax.set_title(f"{cls} {name} ({sid})", fontsize=10)
            ax.set_xticks(range(len(dates)))
            ax.set_xticklabels(dates, rotation=45, ha="right")
            ax.axhline(self.MED, color="#facc15", linestyle="--", linewidth=1.2, alpha=0.9)
            ax.axhline(self.LOW, color="#ef4444", linestyle="--", linewidth=1.2, alpha=0.9)
            ax.text(0.98, self.MED, f"{self.MED}", transform=ax.get_yaxis_transform(),
                    ha="right", va="center", fontsize=8, color="#ca8a04")
            ax.text(0.98, self.LOW, f"{self.LOW}", transform=ax.get_yaxis_transform(),
                    ha="right", va="center", fontsize=8, color="#b91c1c")

            for xi, y in zip(xs, ys):
                if y is not None and not pd.isna(y):
                    ax.annotate(f"{y:.2f}", xy=(xi, y), xytext=(0, 6),
                                textcoords="offset points", ha="center", va="bottom", fontsize=8)

            img_b64 = self._fig_to_base64(fig)
            plt.close(fig)

            card = (
                f'<div class="chart-card" '
                f'data-id="{sid}" '
                f'data-class="{cls}" '
                f'data-has-yellow={"1" if has_yel else "0"} '
                f'data-has-red={"1" if has_red else "0"}>'
                f'<img src="data:image/png;base64,{img_b64}" alt="{cls} {name} ({sid}) GPA chart"/>'
                f'</div>'
            )
            cards.append(card)

        return "\n".join(cards)

    def _build_html(self, data, classes, dates, lang='zh'):
        return self._build_inline_html(data, classes, dates, lang=lang)

    def _build_inline_html(self, data, classes, dates, lang='zh'):
        df = pd.DataFrame(data)
        chart_cards = self._build_chart_cards(df, dates)
        
        # Language labels
        title = 'GPA Analysis Report'
        class_all = '全部班级 All Classes' if lang == 'zh' else 'All Classes'
        red_label = f'红色预警 (<{self.LOW})' if lang == 'zh' else f'Report Alert (<{self.LOW})'
        yel_label = f'黄色预警 ({self.LOW}-{self.MED})' if lang == 'zh' else f'Yellow Alert ({self.LOW}-{self.MED})'
        reset_btn = '重置 Reset' if lang == 'zh' else 'Reset'
        
        # 判断是否显示班级筛选（如果有多个不同的班级）
        has_classes = len(set(c for c in classes if c != 'All')) > 0
        class_select_html = ''
        if has_classes:
            class_select_html = f'''
    <div class="row-center">
      <select id="classSelect">
        <option value="">{class_all}</option>
        {''.join(f'<option value="{c}">{c}</option>' for c in classes if c != 'All')}
      </select>
    </div>'''
        
        html = f"""<!doctype html>
<html lang="{lang}">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>GPA Report</title>
<style>
{self._get_inline_css()}
</style>
</head>
<body>
  <div class="header-bar">
    <h1>{title}</h1>
    <div style="display:flex;gap:10px;align-items:center;">
      <button class="lang-switch" onclick="toggleLang()">EN / 中文</button>
      <button class="export-btn" onclick="exportFilteredPDF()">Export PDF</button>
    </div>
  </div>
  
  <div class="toolbar">
    {class_select_html}
    <div class="row-left">
      <label class="chk">
        <input type="checkbox" id="onlyRed">
        <span class="chip chip-red" data-zh="红色预警 (<3.0)" data-en="Red Alert (<3.0)">红色预警 (&lt;{self.LOW})</span>
      </label>
      <label class="chk">
        <input type="checkbox" id="onlyYellow">
        <span class="chip chip-yellow" data-zh="黄色预警 (3.0-3.5)" data-en="Yellow Alert (3.0-3.5)">黄色预警 (&lt;{self.MED})</span>
      </label>
    </div>
    <div class="row-actions">
      <button class="tab active" data-target="view-table" data-zh="表格 Table" data-en="表格 Table">表格 Table</button>
      <button class="tab" data-target="view-charts" data-zh="折线图 Chart" data-en="折线图 Chart">折线图 Chart</button>
      <button class="tab" data-target="view-stats" data-zh="统计 Stats" data-en="统计 Stats">统计 Stats</button>
      <button id="resetBtn" class="reset" data-zh="重置 Reset" data-en="重置 Reset">重置 Reset</button>
    </div>
  </div>

  <div id="view-table">
    <div class="table-wrap">
      <table id="gpaTable">
        <thead>
          <tr>
            <th class="sticky idx">#</th>
            <th class="sticky name" data-key="姓名">姓名 Name</th>
            <th class="sticky cls" data-key="班级">班级 Class</th>
            <th data-key="变化率" data-zh="变化率" data-en="Change">变化率</th>
            {''.join(f'<th data-key="{d}" data-zh="{d}" data-en="{d}">{d}</th>' for d in dates)}
          </tr>
        </thead>
        <tbody></tbody>
      </table>
    </div>
  </div>

  <div id="view-charts" style="display:none;">
    <div class="charts-grid">{chart_cards}</div>
  </div>

  <div id="view-stats" style="display:none;">
    <div class="table-wrap">
      <table id="statsTable">
        <thead><tr></tr></thead>
        <tbody></tbody>
      </table>
    </div>
  </div>

<script>
const LOW = {self.LOW};
const MED = {self.MED};
const DATES = {json.dumps(dates)};
const ROWS = {json.dumps(data)};
var lang = '{lang}';
{self._get_inline_js()}
</script>
</body>
</html>"""
        return html

    def _get_inline_css(self):
        return """
:root {
    --primary: #6366f1;
    --success: #10b981;
    --danger: #ef4444;
    --warning: #f59e0b;
    --bg: #f8fafc;
    --card-bg: #ffffff;
    --text: #1e293b;
    --text-muted: #64748b;
    --red-bg: #fee2e2;
    --red-fg: #b91c1c;
    --yel-bg: #fef3c7;
    --yel-fg: #b45309;
    --border: #e2e8f0;
}
body {font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;margin:0;padding:24px;background:var(--bg);color:var(--text);}
.header-bar {display:flex;justify-content:space-between;align-items:center;margin-bottom:20px;background:var(--card-bg);padding:16px 24px;border-radius:16px;box-shadow:0 4px 6px rgba(0,0,0,0.07);}
.header-bar h1 {margin:0;font-size:24px;background:linear-gradient(135deg,#6366f1,#8b5cf6);-webkit-background-clip:text;-webkit-text-fill-color:transparent;}
.lang-switch {padding:8px 16px;border:none;background:linear-gradient(135deg,#6366f1,#8b5cf6);color:white;border-radius:8px;cursor:pointer;font-weight:600;font-size:13px;}
.lang-switch:hover {opacity:0.9;}
.export-btn {padding:8px 16px;border:none;background:#10b981;color:white;border-radius:8px;cursor:pointer;font-weight:600;font-size:13px;}
.export-btn:hover {opacity:0.9;}
h1 {text-align:center;font-size:28px;font-weight:700;margin:0 0 20px;background:linear-gradient(135deg,#6366f1,#8b5cf6);-webkit-background-clip:text;-webkit-text-fill-color:transparent;}
.toolbar {background:var(--card-bg);border-radius:16px;padding:20px;margin-bottom:20px;box-shadow:0 4px 6px rgba(0,0,0,0.07);display:grid;grid-template-columns:1fr;gap:12px;}
.row-center {display:flex;justify-content:center;align-items:center;}
.row-left {display:flex;justify-content:flex-start;align-items:center;gap:10px;flex-wrap:wrap;}
.row-actions {display:flex;gap:8px;justify-content:center;align-items:center;flex-wrap:wrap;}
#classSelect {font-size:14px;padding:8px 14px;border:2px solid var(--border);border-radius:10px;min-width:160px;}
.chk {display:flex;align-items:center;gap:6px;cursor:pointer;}
.chip {display:inline-block;padding:6px 12px;border-radius:20px;font-size:13px;font-weight:600;background:#fff;}
.chip-red {color:#dc2626;border:2px solid #fca5a5;background:#fef2f2;}
.chip-yellow {color:#d97706;border:2px solid #fcd34d;background:#fffbeb;}
.tab {padding:10px 20px;border:none;background:#f1f5f9;border-radius:10px;cursor:pointer;font-weight:600;font-size:14px;color:var(--text-muted);transition:all 0.2s;}
.tab:hover {background:#e2e8f0;}
.tab.active {background:linear-gradient(135deg,#6366f1,#8b5cf6);color:white;}
.reset {padding:10px 20px;border:none;background:#fef2f2;border-radius:10px;cursor:pointer;font-weight:600;font-size:14px;color:#dc2626;}
.reset:hover {background:#fee2e2;}
.table-wrap {background:var(--card-bg);border-radius:16px;overflow:hidden;box-shadow:0 4px 6px rgba(0,0,0,0.07);}
table {width:100%;border-collapse:collapse;font-size:13px;}
th,td {padding:12px 14px;border-bottom:1px solid var(--border);text-align:left;}
thead th {background:linear-gradient(180deg,#f8fafc,#f1f5f9);font-weight:600;color:var(--text-muted);position:sticky;top:0;cursor:pointer;}
thead th:hover {background:#e2e8f0;}
th.sticky,td.sticky {position:sticky;background:var(--card-bg);z-index:2;}
th.sticky.idx {left:0;width:60px;text-align:center;}
th.sticky.name {left:60px;width:160px;}
th.sticky.cls {left:220px;width:100px;}
td.sticky.idx {text-align:center;color:var(--text-muted);font-weight:600;}
td.sticky.name {font-weight:600;}
td.sticky.cls {color:var(--text-muted);}
td.red {background:var(--red-bg)!important;color:var(--red-fg)!important;font-weight:700;}
td.yellow {background:var(--yel-bg)!important;color:var(--yel-fg)!important;font-weight:700;}
td.change-pos {color:#059669;font-weight:700;}
td.change-neg {color:#dc2626;font-weight:700;}
.charts-grid {display:grid;grid-template-columns:repeat(auto-fill,minmax(380px,1fr));gap:20px;}
.chart-card {background:var(--card-bg);border-radius:16px;padding:16px;box-shadow:0 4px 6px rgba(0,0,0,0.07);transition:transform 0.2s;}
.chart-card:hover {transform:translateY(-4px);box-shadow:0 8px 12px rgba(0,0,0,0.1);}
.chart-card img {width:100%;border-radius:8px;}
#statsTable {width:100%;}
#statsTable th,#statsTable td {text-align:center;}
#statsTable td.sticky {left:0;width:120px;text-align:left;font-weight:600;}
"""

    def _get_inline_js(self):
        """完整JS脚本"""
        return """
var state = { classFilter: '', onlyYellow: false, onlyRed: false, sortKey: null, sortDir: 'asc' };

function passesFilters(row) {
  if (state.classFilter && row['班级'] !== state.classFilter) return false;
  if (!state.onlyYellow && !state.onlyRed) return true;
  var hasYellow = false, hasRed = false;
  for (var d of DATES) {
    var num = Number(row[d]);
    if (!isNaN(num) && num !== 0) {
      if (num < LOW) hasRed = true;
      else if (num < MED) hasYellow = true;
    }
  }
  if (state.onlyRed && !state.onlyYellow) return hasRed;
  if (state.onlyYellow && !state.onlyRed) return hasYellow;
  if (state.onlyYellow && state.onlyRed) return hasYellow || hasRed;
  return true;
}

function compare(a, b, key) {
  var va = a[key], vb = b[key];
  if (key === '姓名' || key === '班级') {
    return String(va || '').localeCompare(String(vb || ''), 'zh');
  } else {
    if (va == null && vb == null) return 0;
    if (va == null) return 1;
    if (vb == null) return -1;
    return Number(va) - Number(vb);
  }
}

function render() {
  var rows = ROWS.filter(passesFilters);
  
  if (state.sortKey) {
    rows.sort(function(a, b) { return compare(a, b, state.sortKey) * (state.sortDir === 'asc' ? 1 : -1); });
  } else {
    rows.sort(function(a, b) {
      var c = String(a['班级'] || '').localeCompare(String(b['班级'] || ''), 'zh');
      return c !== 0 ? c : String(a['姓名'] || '').localeCompare(String(b['姓名'] || ''), 'zh');
    });
  }
  
  var tbody = document.querySelector('#gpaTable tbody');
  tbody.innerHTML = '';
  
  rows.forEach(function(r, i) {
    var tr = document.createElement('tr');
    
    var tdIdx = document.createElement('td');
    tdIdx.className = 'sticky idx';
    tdIdx.textContent = i + 1;
    tr.appendChild(tdIdx);
    
    var tdName = document.createElement('td');
    tdName.className = 'sticky name';
    tdName.textContent = r['姓名'] || '';
    tr.appendChild(tdName);
    
    var tdClass = document.createElement('td');
    tdClass.className = 'sticky cls';
    tdClass.textContent = r['班级'] || '';
    tr.appendChild(tdClass);
    
    var tdCR = document.createElement('td');
    var cr = r['变化率'];
    if (cr !== undefined && cr !== null && !isNaN(cr)) {
      var pct = (cr * 100).toFixed(1) + '%';
      tdCR.textContent = (cr > 0 ? '+' : '') + pct;
      tdCR.className = cr > 0 ? 'change-pos' : 'change-neg';
    } else {
      tdCR.textContent = '—';
    }
    tr.appendChild(tdCR);
    
    DATES.forEach(function(d) {
      var td = document.createElement('td');
      var num = Number(r[d]);
      if (!isNaN(num)) {
        td.textContent = num.toFixed(2);
        if (num < LOW) td.classList.add('red');
        else if (num < MED) td.classList.add('yellow');
      } else {
        td.textContent = r[d] || '';
      }
      tr.appendChild(td);
    });
    
    tbody.appendChild(tr);
  });
}

function updateChartsVisibility() {
  document.querySelectorAll('.chart-card').forEach(function(card) {
    var cls = card.getAttribute('data-class') || '';
    var hasYellow = card.getAttribute('data-has-yellow') === '1';
    var hasRed = card.getAttribute('data-has-red') === '1';
    
    var ok = true;
    if (state.classFilter && cls !== state.classFilter) ok = false;
    if (state.onlyRed && !hasRed) ok = false;
    if (state.onlyYellow && !hasYellow) ok = false;
    
    card.style.display = ok ? '' : 'none';
  });
}

function computeStatsAll() {
  // 使用筛选后的数据计算统计
  var filteredRows = ROWS.filter(passesFilters);
  
  var avg = new Array(DATES.length).fill(null);
  var yel = new Array(DATES.length).fill(0);
  var red = new Array(DATES.length).fill(0);
  var above4 = new Array(DATES.length).fill(0);
  
  for (var j = 0; j < DATES.length; j++) {
    var d = DATES[j];
    var sum = 0, cnt = 0;
    
    for (var r of filteredRows) {
      var v = Number(r[d]);
      if (!isNaN(v) && v > 0) { sum += v; cnt += 1; }
      if (!isNaN(v) && v > 0) {
        if (v >= 4.0) above4[j] += 1;
        else if (v < LOW) red[j] += 1;
        else if (v < MED) yel[j] += 1;
      }
    }
    
    avg[j] = (cnt > 0) ? (sum / cnt) : null;
  }
  return { avg: avg, yel: yel, red: red, above4: above4 };
}

function renderStatsAll() {
  var table = document.getElementById('statsTable');
  if (!table) return;
  
  var trh = table.querySelector('thead tr');
  trh.innerHTML = '<th data-zh="指标" data-en="Metric">指标</th>' + DATES.map(function(d) { return '<th>' + d + '</th>'; }).join('');
  
  var S = computeStatsAll();
  
  var rows = [
    { name: '平均分', nameEn: 'Average', vals: S.avg.map(function(v) { return v == null ? '—' : v.toFixed(2); }) },
    { name: '>=4.0', nameEn: '>=4.0', vals: S.above4.map(function(n) { return String(n); }) },
    { name: '黄色预警', nameEn: 'Yellow Alert', vals: S.yel.map(function(n) { return String(n); }) },
    { name: '红色预警', nameEn: 'Red Alert', vals: S.red.map(function(n) { return String(n); }) }
  ];
  
  var tbody = table.querySelector('tbody');
  tbody.innerHTML = '';
  
  rows.forEach(function(row) {
    var tr = document.createElement('tr');
    var name = lang === 'zh' ? row.name : row.nameEn;
    tr.innerHTML = '<td class="sticky">' + name + '</td>' + row.vals.map(function(txt) { return '<td>' + txt + '</td>'; }).join('');
    tbody.appendChild(tr);
  });
}

function bindUI() {
  document.getElementById('classSelect').addEventListener('change', function() {
    state.classFilter = this.value;
    render();
    updateChartsVisibility();
    renderStatsAll();
  });
  
  document.getElementById('onlyYellow').addEventListener('change', function() {
    state.onlyYellow = this.checked;
    render();
    updateChartsVisibility();
    renderStatsAll();
  });
  
  document.getElementById('onlyRed').addEventListener('change', function() {
    state.onlyRed = this.checked;
    render();
    updateChartsVisibility();
    renderStatsAll();
  });
  
  document.getElementById('resetBtn').addEventListener('click', function() {
    document.getElementById('classSelect').value = '';
    document.getElementById('onlyYellow').checked = false;
    document.getElementById('onlyRed').checked = false;
    state = { classFilter: '', onlyYellow: false, onlyRed: false, sortKey: null, sortDir: 'asc' };
    render();
    updateChartsVisibility();
    renderStatsAll();
  });
  
  document.querySelectorAll('.tab').forEach(function(btn) {
    btn.addEventListener('click', function() {
      document.querySelectorAll('.tab').forEach(function(b) { b.classList.remove('active'); });
      btn.classList.add('active');
      var target = btn.getAttribute('data-target');
      document.getElementById('view-table').style.display = (target === 'view-table') ? '' : 'none';
      document.getElementById('view-charts').style.display = (target === 'view-charts') ? '' : 'none';
      document.getElementById('view-stats').style.display = (target === 'view-stats') ? '' : 'none';
      updateChartsVisibility();
      if (target === 'view-stats') renderStatsAll();
    });
  });
  
  document.querySelector('#gpaTable thead').addEventListener('click', function(e) {
    var th = e.target.closest('th');
    if (!th) return;
    var key = th.getAttribute('data-key');
    if (!key) return;
    if (state.sortKey === key) state.sortDir = (state.sortDir === 'asc') ? 'desc' : 'asc';
    else { state.sortKey = key; state.sortDir = 'asc'; }
    render();
  });
}

document.addEventListener('DOMContentLoaded', function() {
  bindUI();
  render();
  updateChartsVisibility();
});

function toggleLang() {
  lang = lang === 'zh' ? 'en' : 'zh';
  document.querySelectorAll('[data-zh]').forEach(function(el) {
    var zh = el.getAttribute('data-zh');
    var en = el.getAttribute('data-en');
    el.textContent = lang === 'zh' ? zh : en;
  });
  render();
  renderStatsAll();
}

async function exportFilteredPDF() {
  // 获取当前筛选后的数据
  var filteredRows = ROWS.filter(passesFilters);
  
  if (filteredRows.length === 0) {
    alert(lang === 'zh' ? '没有数据可导出' : 'No data to export');
    return;
  }
  
  // 获取班级列表
  var classes = [...new Set(filteredRows.map(function(r) { return r['班级']; }))];
  
  // 调用后端API
  try {
    var btn = document.querySelector('.export-btn');
    var originalText = btn.textContent;
    btn.textContent = lang === 'zh' ? '导出中...' : 'Exporting...';
    btn.disabled = true;
    
    var response = await fetch('/export-pdf-filtered', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({
        rows: filteredRows,
        classes: classes,
        dates: DATES,
        lang: lang
      })
    });
    
    var result = await response.json();
    
    if (result.success) {
      // 下载PDF
      window.location.href = result.download_url;
    } else {
      alert(lang === 'zh' ? '导出失败: ' + result.error : 'Export failed: ' + result.error);
    }
  } catch (error) {
    alert(lang === 'zh' ? '导出失败: ' + error.message : 'Export failed: ' + error.message);
  } finally {
    btn.disabled = false;
    btn.textContent = originalText;
  }
}
"""

    def _add_change_rate(self, rows, dates):
        if len(dates) < 2:
            for r in rows:
                r["变化率"] = None
            return rows

        last, prev = dates[-1], dates[-2]
        for r in rows:
            cur = r.get(str(last))
            pre = r.get(str(prev))
            try:
                cur = float(cur) if cur is not None else None
                pre = float(pre) if pre is not None else None
            except:
                cur = pre = None

            if pre in (None, 0) or cur is None:
                r["变化率"] = None
            else:
                r["变化率"] = (cur - pre) / pre
        return rows
